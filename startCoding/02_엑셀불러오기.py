import openpyxl

fpath = r'C:\Users\82104\OneDrive\바탕 화면\Workspace\my-crawling-test\sample.xlsx'

# 1) 엑셀 불러오기
wb = openpyxl.load_workbook(fpath)

# 2) 시트 선택
ws = wb['테스트 시트']

# 3) 데이터 수정
ws['A3'] = 456
ws['B3'] = '이부자'

# 4) 엑셀 저장
wb.save(fpath)
